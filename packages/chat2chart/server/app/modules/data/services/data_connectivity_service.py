"""
Universal Data Connectivity Service
Handles file uploads, database connections, and data source management
"""

import logging
import os
import pandas as pd
import json
import re
import sqlalchemy as sa
from typing import Dict, List, Optional, Any, Union
from datetime import datetime, date
import tempfile
from pathlib import Path
from .database_connector_service import DatabaseConnectorService
from app.modules.data.services.ai_schema_service import AISchemaService
from app.db.session import async_operation_lock
from app.modules.data.utils.credentials import encrypt_credentials, decrypt_credentials

logger = logging.getLogger(__name__)


class DataConnectivityService:
    """Service for handling data connectivity and file uploads"""
    
    def __init__(self):
        self.max_file_size = 50 * 1024 * 1024  # 50MB
        self.supported_formats = ['csv', 'xlsx', 'xls', 'json', 'tsv', 'parquet', 'parq', 'snappy']
        
        # Enhanced file processing options
        self.file_processing_configs = {
            'csv': {
                'delimiters': [',', ';', '\t', '|', ' '],
                'encodings': ['utf-8', 'latin-1', 'cp1252'],
                'auto_detect': True
            },
            'tsv': {
                'delimiters': ['\t'],
                'encodings': ['utf-8', 'latin-1'],
                'auto_detect': False
            },
            'xlsx': {
                'sheet_selection': 'auto',
                'header_row': 0,
                'skip_rows': 0
            },
            'xls': {
                'sheet_selection': 'auto',
                'header_row': 0,
                'skip_rows': 0
            },
            'json': {
                'flatten_nested': True,
                'max_nesting_level': 3
            },
            'parquet': {
                'columns': 'auto',
                'partitions': 'auto'
            }
        }
        
        # Data source registry (in production, this would be in database)
        self.data_sources = {}
        
        # Initialize demo data for testing
        self._initialize_demo_data()
        
        # Initialize database connector service
        self.database_connector = DatabaseConnectorService()
        self.ai_schema_service = AISchemaService()  # Add AI schema service
    
    def _make_json_serializable(self, obj: Any) -> Any:
        """Recursively convert date/datetime objects to ISO format strings for JSON serialization"""
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {key: self._make_json_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        elif pd.isna(obj):
            return None
        else:
            return obj

    async def test_database_connection(self, connection_request: Dict[str, Any]) -> Dict[str, Any]:
        """Test database connection without storing credentials"""
        try:
            logger.info(f"🔌 Testing database connection: {connection_request.get('type')}")
            
            db_type = connection_request.get('type', '').lower()
            
            # Get supported databases from DatabaseConnectorService
            supported_dbs = self.database_connector.get_supported_databases()
            
            if db_type not in supported_dbs:
                return {
                    'success': False,
                    'error': f'Unsupported database type: {db_type}. Supported: {supported_dbs}'
                }
            
            # Test connection using DatabaseConnectorService
            try:
                test_result = await self.database_connector.test_connection(connection_request)
                return test_result
                    
            except Exception as e:
                logger.error(f"❌ Database connection test error: {str(e)}")
                return {
                    'success': False,
                    'error': f'Connection test failed: {str(e)}'
                }
                
        except Exception as e:
            logger.error(f"❌ Database connection test failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def store_database_connection(self, connection_request: Dict[str, Any], user_id: str) -> Dict[str, Any]:
        """Store database connection configuration
        
        NOTE: This method validates the connection BEFORE encrypting credentials.
        This ensures the validation works with plain credentials, and also provides
        safety for users who skip the test step and directly click save.
        
        Args:
            connection_request: Database connection configuration
            user_id: User ID (required for data isolation and security)
        """
        if not user_id:
            raise ValueError("user_id is required for database connections")
        
        try:
            db_type = str(connection_request.get('type', '')).lower()
            logger.info(f"💾 Storing database connection of type '{db_type}' for user {user_id}")

            # ALWAYS validate the connection first with PLAIN credentials.
            # This prevents persisting phantom data sources when credentials or networking are invalid,
            # and provides safety for users who skip the test step.
            try:
                test_result = await self.database_connector.test_connection(connection_request)
                if not test_result.get('success'):
                    error_msg = test_result.get('error') or 'Database connection test failed'
                    logger.warning(f"❌ Connection validation failed; not storing data source: {error_msg}")
                    return {
                        'success': False,
                        'error': error_msg
                    }
            except Exception as validation_error:
                logger.error(f"❌ Unexpected error while validating connection before store: {validation_error}")
                return {
                    'success': False,
                    'error': f'Connection validation failed: {validation_error}'
                }

            # Generate unique ID for the connection
            connection_id = f"db_{connection_request.get('type')}_{int(datetime.now().timestamp())}"
            
            # Prepare in-memory representation upfront to avoid unbound variable
            connection_data = {
                'id': connection_id,
                'type': 'database',
                'db_type': connection_request.get('type'),
                'name': connection_request.get('name', f"{connection_request.get('type')}_connection"),
                'config': connection_request,
                'created_at': datetime.now().isoformat(),
                'status': 'connected',
                'connection_type': 'database'
            }
            
            # Store connection info in database
            try:
                from app.db.models import DataSource
                from app.db.session import async_session

                async with async_session() as db:
                    # Create new data source record
                    # Ensure any sensitive credentials are stored encrypted in DB
                    try:
                        from app.modules.data.utils.credentials import encrypt_credentials
                        safe_config = encrypt_credentials(connection_request)
                        # Log encryption status for debugging
                        if safe_config != connection_request:
                            logger.info(f"✅ Credentials encrypted for {connection_request.get('type')} connection")
                        else:
                            logger.warning(f"⚠️ Credentials not encrypted (ENCRYPTION_KEY may not be set) for {connection_request.get('type')} connection")
                    except Exception as encrypt_error:
                        logger.error(f"❌ Failed to encrypt credentials: {encrypt_error}")
                        safe_config = connection_request
                    
                    new_source = DataSource(
                        id=connection_id,
                        name=connection_request.get('name') or f"{connection_request.get('type')}_connection",
                        type='database',
                        format=connection_request.get('type'),
                        db_type=connection_request.get('type'),
                        size=0,  # Database connections don't have file size
                        row_count=0,  # Will be populated when schema is fetched
                        schema=json.dumps({
                    'type': connection_request.get('type'),
                    'host': connection_request.get('host'),
                    'port': connection_request.get('port'),
                    'database': connection_request.get('database'),
                    'username': connection_request.get('username'),
                            'ssl_mode': connection_request.get('ssl_mode', 'prefer'),
                            'connection_string': connection_request.get('uri'),
                            'encrypt': connection_request.get('encrypt', False)
                        }),
                        connection_config=json.dumps(safe_config),
                        metadata=json.dumps({
                            'connection_type': 'database',
                            'status': 'connected',
                            'created_at': datetime.now().isoformat()
                        }),
                        user_id=user_id,  # Required - validated at function entry
                        is_active=True,
                        created_at=datetime.now(),
                        updated_at=datetime.now(),
                        last_accessed=datetime.now()
                    )
                    
                    db.add(new_source)
                    await db.commit()
                    await db.refresh(new_source)
                    
                    logger.info(f"✅ Database connection saved to database: {connection_id}")
                    # Keep in-memory registry aligned
                    connection_data.update({
                        'name': new_source.name,
                        'db_type': new_source.db_type
                    })
            except Exception as db_error:
                logger.error(f"❌ Failed to save to database: {str(db_error)}")
                # Fallback to memory storage (connection_data already prepared)
            self.data_sources[connection_id] = connection_data
            
            # Create connection in DatabaseConnectorService for future queries
            try:
                conn_result = await self.database_connector.create_connection(connection_request)
                if conn_result.get('success'):
                    logger.info(f"✅ Database connection engine created: {connection_id}")
            except Exception as conn_error:
                logger.warning(f"⚠️ Connection engine creation failed (queries may not work): {str(conn_error)}")
            
            logger.info(f"✅ Database connection stored: {connection_id}")
            
            return {
                'success': True,
                'data_source_id': connection_id,
                'connection_info': {
                    'id': connection_id,
                    'type': 'database',
                    'db_type': connection_request.get('type'),
                    'name': connection_request.get('name', f"{connection_request.get('type')}_connection"),
                    'status': 'connected'
                }
            }
            
        except Exception as e:
            logger.error(f"❌ Failed to store database connection: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def _read_file_data(self, file_path: str, file_format: str, limit: int = 100) -> List[Dict[str, Any]]:
        """Read data from a file based on its format"""
        try:
            logger.info(f"📁 Reading {file_format} file: {file_path}")
            
            if file_format == 'csv':
                import pandas as pd
                df = pd.read_csv(file_path, nrows=limit)
                return df.to_dict('records')
            elif file_format == 'xlsx':
                import pandas as pd
                df = pd.read_excel(file_path, nrows=limit)
                return df.to_dict('records')
            elif file_format == 'json':
                import json
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, list):
                    return data[:limit]
                else:
                    return [data]
            elif file_format == 'parquet':
                import pandas as pd
                df = pd.read_parquet(file_path)
                return df.head(limit).to_dict('records')
            else:
                logger.warning(f"⚠️ Unsupported file format: {file_format}")
                return []
                
        except Exception as e:
            logger.error(f"❌ Failed to read file {file_path}: {str(e)}")
            return []
    
    async def execute_query_on_source(self, source_id: str, query: str) -> Dict[str, Any]:
        """Execute a custom query on a data source"""
        try:
            logger.info(f"🔍 Executing query on source: {source_id}")
            logger.info(f"🔍 Query: {query}")
            
            # Get the data source
            source = await self.get_data_source_by_id(source_id)
            if not source:
                return {
                    'success': False,
                    'error': f'Data source {source_id} not found'
                }
            
            # For demo data sources, implement basic query parsing
            if source.get('source') == 'demo_data' or source.get('id', '').startswith('demo_'):
                data = source.get('sample_data', [])
                
                # Simple query parsing for demo purposes
                if 'count' in query.lower():
                    # Count query
                    result = len(data)
                    return {
                        'success': True,
                        'data': [{'count': result}],
                        'total_rows': 1,
                        'query_type': 'count'
                    }
                elif 'select' in query.lower() and 'from' in query.lower():
                    # Basic SELECT query simulation
                    # This is a simplified implementation for demo purposes
                    try:
                        # Parse basic SELECT * FROM table
                        if 'select *' in query.lower():
                            return {
                                'success': True,
                                'data': data,
                                'total_rows': len(data),
                                'query_type': 'select_all'
                            }
                        else:
                            # For other queries, return sample data
                            return {
                                'success': True,
                                'data': data[:10],  # Limit to 10 rows
                                'total_rows': len(data[:10]),
                                'query_type': 'select_limited'
                            }
                    except Exception as parse_error:
                        logger.warning(f"Query parsing failed, returning sample data: {parse_error}")
                        return {
                            'success': True,
                            'data': data[:10],
                            'total_rows': len(data[:10]),
                            'query_type': 'fallback'
                        }
                else:
                    # Default: return sample data
                    return {
                        'success': True,
                        'data': data,
                        'total_rows': len(data),
                        'query_type': 'default'
                    }
            
            # For file sources, this would parse and execute the query
            elif source.get('type') == 'file':
                return {
                    'success': False,
                    'error': 'File query execution not yet implemented'
                }
            
            # For database sources, this would execute SQL
            elif source.get('type') == 'database':
                return {
                    'success': False,
                    'error': 'Database query execution not yet implemented'
                }
            
            else:
                return {
                    'success': False,
                    'error': f'Unsupported data source type: {source.get("type")}'
                }
                
        except Exception as e:
            logger.error(f"❌ Failed to execute query on source {source_id}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_source_schema(self, source_id: str) -> Dict[str, Any]:
        """Get schema information for a specific data source"""
        try:
            logger.info(f"🔍 Getting schema for source: {source_id}")
            
            # Get the data source
            source = await self.get_data_source_by_id(source_id)
            if not source:
                return {
                    'success': False,
                    'error': f'Data source {source_id} not found'
                }
            
            # Return schema information
            schema = source.get('schema', {})
            # Parse JSON string schemas from database if necessary
            if isinstance(schema, str):
                try:
                    schema = json.loads(schema)
                except json.JSONDecodeError:
                    logger.warning("⚠️ Stored schema is a string but not valid JSON; returning empty schema")
                    schema = {}
            if schema:
                logger.info(f"✅ Returning schema for source {source_id}")
                return {
                    'success': True,
                    'schema': schema,
                    'source_id': source_id,
                    'source_name': source.get('name', 'Unknown'),
                    'source_type': source.get('type', 'Unknown')
                }
            else:
                return {
                    'success': False,
                    'error': f'No schema available for data source {source_id}'
                }
                
        except Exception as e:
            logger.error(f"❌ Failed to get schema for source {source_id}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_data_from_source(self, source_id: str, limit: int = 100) -> Dict[str, Any]:
        """Get data from a specific data source"""
        try:
            logger.info(f"🔍 Getting data from source: {source_id}")
            
            # Get the data source
            source = await self.get_data_source_by_id(source_id)
            if not source:
                return {
                    'success': False,
                    'error': f'Data source {source_id} not found'
                }
            
            # For demo data sources, return sample data
            if source.get('source') == 'demo_data' or source.get('id', '').startswith('demo_'):
                sample_data = source.get('sample_data', [])
                logger.info(f"✅ Returning {len(sample_data)} sample data rows from demo source")
                return {
                    'success': True,
                    'data': sample_data,
                    'total_rows': len(sample_data),
                    'source': source
                }
            
            # For file sources, read the actual file
            elif source.get('type') == 'file':
                file_path = source.get('file_path')
                if not file_path:
                    return {
                        'success': False,
                        'error': 'No file path available for this data source'
                    }
                
                # Read file data based on format
                file_format = source.get('format', 'csv').lower()
                data = await self._read_file_data(file_path, file_format, limit)
                
                return {
                    'success': True,
                    'data': data,
                    'total_rows': len(data),
                    'source': source
                }
            
            # For database sources, this would query the database
            elif source.get('type') == 'database':
                return {
                    'success': False,
                    'error': 'Database data retrieval not yet implemented'
                }
            
            else:
                return {
                    'success': False,
                    'error': f'Unsupported data source type: {source.get("type")}'
                }
                
        except Exception as e:
            logger.error(f"❌ Failed to get data from source {source_id}: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def get_data_source_by_id(self, source_id: str) -> Optional[Dict[str, Any]]:
        """Get a specific data source by ID"""
        try:
            # Development aliases: map common demo UI ids to available demo sources
            alias_map = {
                'duckdb_local': 'demo_sales_data',
                'csv_sales': 'demo_sales_data',
                'snowflake_warehouse': 'demo_customers_data',
                'postgresql_prod': 'demo_customers_data',
            }
            source_id = alias_map.get(source_id, source_id)
            # First check in-memory demo sources
            if source_id in self.data_sources:
                logger.info(f"✅ Found data source {source_id} in demo sources")
                return self.data_sources[source_id]
            
            # Then check database
            from app.modules.data.models import DataSource
            from app.db.session import async_session

            # Use async_session() which returns a session that can be used as context manager
            from app.db.session import async_session
            async with async_session() as db:
                # Ensure a per-session operation lock exists when the caller used
                # the async_session factory directly (not via get_async_session).
                # This prevents asyncpg 'another operation is in progress' errors
                # by serializing DB operations within this session.
                try:
                    if not getattr(db, '_op_lock', None):
                        import asyncio as _asyncio

                        db._op_lock = _asyncio.Lock()
                except Exception:
                    db._op_lock = None
                from sqlalchemy import select
                query = select(DataSource).where(DataSource.id == source_id, DataSource.is_active == True)
                result = await db.execute(query)
                source = result.scalar_one_or_none()
                
                if source:
                    source_dict = {
                        'id': source.id,
                        'name': source.name,
                        'type': source.type,
                        'format': source.format,
                        # expose persisted inline/sample data and file path if present
                        'sample_data': getattr(source, 'sample_data', None),
                        'file_path': getattr(source, 'file_path', None),
                        'db_type': source.db_type,
                        'size': source.size,
                        'row_count': source.row_count,
                        'schema': source.schema,
                        'user_id': getattr(source, 'user_id', None),
                        'created_at': source.created_at.isoformat() if source.created_at else None,
                        'updated_at': source.updated_at.isoformat() if source.updated_at else None,
                        'is_active': source.is_active,
                        'last_accessed': source.last_accessed.isoformat() if source.last_accessed else None
                    }
                    
                    # CRITICAL: Include connection_config for database/warehouse sources and decrypt credentials
                    if (source.type == 'database' or source.type == 'warehouse') and source.connection_config:
                        try:
                            import json
                            config = json.loads(source.connection_config) if isinstance(source.connection_config, str) else source.connection_config
                            
                            # CRITICAL: Decrypt credentials before returning
                            try:
                                from app.modules.data.utils.credentials import decrypt_credentials
                                config = decrypt_credentials(config)
                                logger.debug(f"✅ Decrypted credentials for data source {source_id}")
                            except Exception as decrypt_error:
                                logger.warning(f"⚠️ Could not decrypt credentials for {source_id} (may not be encrypted): {decrypt_error}")
                            
                            # Add to multiple keys for compatibility with query engine
                            source_dict['connection_config'] = config
                            source_dict['connection_info'] = config
                            source_dict['config'] = config
                            source_dict['metadata'] = config  # Also add as metadata for compatibility
                        except Exception as config_error:
                            logger.error(f"❌ Error parsing connection_config for {source_id}: {config_error}")
                            source_dict['connection_config'] = {}
                            source_dict['connection_info'] = {}
                    
                    logger.info(f"✅ Found data source {source_id} in database (type: {source.type}, has_config: {bool(source_dict.get('connection_config'))})")
                    return source_dict
                
                logger.warning(f"⚠️ Data source {source_id} not found in database or demo sources")
                return None
                
        except Exception as e:
            logger.error(f"❌ Failed to get data source {source_id}: {str(e)}")
            # Fallback to in-memory sources
            return self.data_sources.get(source_id)
    
    async def get_data_sources(self, offset: int = 0, limit: int = 100) -> List[Dict[str, Any]]:
        """Get all data sources with pagination from database"""
        try:
            logger.info(f"🔍 Getting data sources (offset: {offset}, limit: {limit})")
            logger.info(f"🔍 Available demo sources: {list(self.data_sources.keys())}")
            
            from app.modules.data.models import DataSource
            from app.db.session import async_session

            async with async_session() as db:
                # Query database for data sources using async session
                from sqlalchemy import select, func
                
                # Count total
                count_query = select(func.count(DataSource.id)).where(DataSource.is_active == True)
                total_result = await db.execute(count_query)
                total = total_result.scalar()
                
                # Get paginated results
                query = select(DataSource).where(DataSource.is_active == True).offset(offset).limit(limit)
                result = await db.execute(query)
                sources = result.scalars().all()
                
                # Convert to dictionary format
                result_list = []
                for source in sources:
                    source_dict = {
                        'id': source.id,
                        'name': source.name,
                        'type': source.type,
                        'format': source.format,
                        'db_type': source.db_type,
                        'size': source.size,
                        'row_count': source.row_count,
                        'schema': source.schema,
                        'created_at': source.created_at.isoformat() if source.created_at else None,
                        'updated_at': source.updated_at.isoformat() if source.updated_at else None,
                        'is_active': source.is_active,
                        'last_accessed': source.last_accessed.isoformat() if source.last_accessed else None
                    }
                    result_list.append(source_dict)
                
                logger.info(f"✅ Retrieved {len(result_list)} data sources from database")
                
                # Always include demo data sources for testing
                demo_sources = list(self.data_sources.values())
                logger.info(f"🔍 Adding {len(demo_sources)} demo sources")
                all_sources = result_list + demo_sources
                logger.info(f"🔍 Total sources after combining: {len(all_sources)}")
                
                # Apply pagination to combined sources
                result = all_sources[offset:offset + limit]
                logger.info(f"🔍 Returning {len(result)} sources after pagination")
                return result
                
        except Exception as e:
            logger.error(f"❌ Failed to get data sources from database: {str(e)}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            # Fallback to in-memory sources
            sources = list(self.data_sources.values())
            logger.info(f"✅ Using {len(sources)} in-memory demo data sources")
            return sources[offset:offset + limit]

    async def _save_data_source_to_db(self, data_source: Dict[str, Any]) -> bool:
        """Save data source to database"""
        try:
            from app.modules.data.models import DataSource
            from app.db.session import async_session

            async with async_session() as db:
                # Check if data source already exists
                from sqlalchemy import select
                existing_query = select(DataSource).where(DataSource.id == data_source['id'])
                existing_result = await db.execute(existing_query)
                existing = existing_result.scalar_one_or_none()
                
                if existing:
                    # Update existing
                    existing.name = data_source['name']
                    existing.type = data_source['type']
                    existing.format = data_source.get('format')
                    existing.size = data_source.get('size')
                    existing.row_count = data_source.get('row_count')
                    existing.schema = data_source.get('schema')
                    existing.file_path = data_source.get('file_path')  # Now stores object_key
                    existing.original_filename = data_source.get('original_filename')
                    existing.sample_data = data_source.get('sample_data')  # Save sample_data as JSON
                    from datetime import timezone
                    existing.updated_at = datetime.now(timezone.utc)
                    logger.info(f"✅ Updated data source {data_source['id']} in database.")
                else:
                    # Create new
                    # Use user_id from data_source dict (passed from options)
                    user_id = data_source.get('user_id')
                    
                    if not user_id:
                        raise ValueError(f"user_id is required when creating data source {data_source.get('id')}. Ensure user_id is passed in options.")
                    
                    new_data_source = DataSource(
                        id=data_source['id'],
                        name=data_source['name'],
                        type=data_source['type'],
                        format=data_source.get('format'),
                        size=data_source.get('size'),
                        row_count=data_source.get('row_count'),
                        schema=data_source.get('schema'),
                        file_path=data_source.get('file_path'),  # Now stores object_key
                        original_filename=data_source.get('original_filename'),
                        sample_data=data_source.get('sample_data'),  # Save sample_data as JSON
                        user_id=user_id  # Required - validated above
                    )
                    db.add(new_data_source)
                    logger.info(f"✅ Saved new data source {data_source['id']} to database (user_id={user_id}).")
                
                await db.commit()
                return True
                
        except Exception as e:
            logger.error(f"❌ Failed to save data source to database: {str(e)}", exc_info=True)
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return False

    async def process_uploaded_file(
        self, 
        file_path: str,  # Temp file path for processing
        original_filename: str, 
        options: Optional[Dict[str, Any]] = None,
        object_key: Optional[str] = None  # NEW: Object key from PostgreSQL storage
    ) -> Dict[str, Any]:
        """Process uploaded file and extract data"""
        try:
            logger.info(f"📁 Processing uploaded file: {original_filename}")
            
            if options is None:
                options = {}
            
            file_extension = self._get_file_extension(original_filename)
            
            if file_extension not in self.supported_formats:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Process file based on extension
            if file_extension == 'csv':
                data, schema = await self._process_csv_file(file_path, options.get('delimiter', ','), options.get('encoding', 'utf-8'))
            elif file_extension == 'tsv':
                data, schema = await self._process_csv_file(file_path, '\t', 'utf-8')
            elif file_extension == 'parquet':
                data, schema = await self._process_parquet_file(file_path)
            elif file_extension == 'json':
                data, schema = await self._process_json_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Enhance schema with AI insights (skip for preview-only or if disabled for performance)
            preview_only = options.get('preview_only', False)
            skip_ai_enhancement = preview_only or options.get('skip_ai_enhancement', False)
            
            if skip_ai_enhancement:
                # Use basic schema without AI enhancement for faster processing
                enhanced_schema = schema
                logger.info("⏩ Skipping AI schema enhancement for faster processing")
            else:
                # Only enhance schema for a sample of data to improve performance
                sample_size = min(100, len(data))  # Use first 100 rows for schema enhancement
                sample_data = data[:sample_size] if data else []
                enhanced_schema = await self.ai_schema_service.generate_enhanced_schema(
                    sample_data, schema, original_filename, "file"
                )
                logger.info(f"✅ AI schema enhancement completed using {sample_size} sample rows")
            
            # Generate data source metadata
            user_id = options.get('user_id') if options else None
            name = options.get('name') if options else original_filename
            
            # Get file size from temp file
            file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            
            data_source = {
                'id': f"file_{int(datetime.now().timestamp())}",
                'name': name or original_filename,
                'type': 'file',
                'format': file_extension,
                'size': file_size,
                'uploaded_at': datetime.now().isoformat(),
                'schema': enhanced_schema,
                'row_count': len(data),
                'file_path': object_key,  # NEW: Store object_key, not file path
                'original_filename': original_filename,
                'preview': data[:10] if len(data) > 10 else data,  # First 10 rows for preview
                # Add fields expected by frontend
                'uuid_filename': f"file_{int(datetime.now().timestamp())}_{original_filename}",
                'content_type': f"application/{file_extension}",
                'storage_type': 'postgresql',  # Updated: now using PostgreSQL storage
                'user_id': user_id  # Pass user_id from options
            }
            
            # Conditional in-memory storage based on upload_with_prompt flag
            upload_with_prompt = options.get('upload_with_prompt', False)
            max_sample_rows = 10000  # Increased from 1000
            
            if upload_with_prompt:
                # Store in memory for immediate query processing
                if len(data) > max_sample_rows:
                    data_source['sample_data'] = data[:max_sample_rows]
                    data_source['data'] = data[:max_sample_rows]
                else:
                    data_source['data'] = data
                    data_source['sample_data'] = data
                
                # Store in memory cache
                self.data_sources[data_source['id']] = data_source
                logger.info(f"💾 Stored in memory for immediate processing")
            else:
                # Don't store in memory - only save to DB
                data_source['sample_data'] = data[:max_sample_rows] if len(data) > max_sample_rows else data
                logger.info(f"⏩ Skipping in-memory storage (file uploaded without prompt)")
            
            # Always save to database
            if not preview_only:
                save_success = await self._save_data_source_to_db(data_source)
                if not save_success:
                    logger.error(f"❌ Failed to save data source {data_source.get('id')} to database")
                    raise Exception(f"Failed to save data source to database. Check logs for details.")
            
            logger.info(f"✅ File processed successfully: {len(data)} rows, {len(enhanced_schema['columns'])} columns")
            
            return {
                'success': True,
                'data_source': data_source,
                'data': data if options.get('include_data') else None
            }
            
        except Exception as error:
            logger.error(f"❌ File processing failed: {str(error)}")
            
            return {
                'success': False,
                'error': str(error)
            }

    async def upload_file(
        self,
        file_content: bytes,
        filename: str,
        options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Upload and process a file from content"""
        tmp_file_path = None
        try:
            logger.info(f"📁 File upload request: {filename}")
            
            if options is None:
                options = {}
            
            # Validate file size
            if len(file_content) > self.max_file_size:
                raise ValueError(f"File too large. Maximum size: {self.max_file_size / (1024*1024):.1f}MB")
            
            # Validate file format
            file_extension = self._get_file_extension(filename)
            if file_extension not in self.supported_formats:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Get user_id from options
            user_id = options.get('user_id')
            if not user_id:
                raise ValueError("user_id is required for file upload")
            
            # Store file in PostgreSQL
            from app.modules.data.services.postgres_storage_service import PostgresStorageService
            storage_service = PostgresStorageService()
            
            content_type = f"application/{file_extension}"
            
            # Store file in PostgreSQL and get object_key
            object_key = await storage_service.store_file(
                file_content=file_content,
                user_id=user_id,
                original_filename=filename,
                content_type=content_type
            )
            
            logger.info(f"💾 File stored in PostgreSQL: {object_key}")
            
            # Create temp file for processing (will be deleted after processing)
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_extension}") as tmp_file:
                tmp_file.write(file_content)
                tmp_file_path = tmp_file.name
            
            # Process the uploaded file (pass object_key)
            result = await self.process_uploaded_file(tmp_file_path, filename, options, object_key)
            
            if result['success']:
                logger.info(f"✅ File upload completed successfully: {filename}")
                return result
            else:
                return result
                
        except Exception as error:
            logger.error(f"❌ File upload failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }
        finally:
            # Clean up temp file
            if tmp_file_path and os.path.exists(tmp_file_path):
                try:
                    os.unlink(tmp_file_path)
                except Exception as e:
                    logger.warning(f"⚠️ Failed to clean up temp file: {str(e)}")

    async def _process_csv_file(self, file_path: str, delimiter: str = ',', encoding: str = 'utf-8') -> tuple:
        """Process CSV/TSV files using DuckDB for fast, direct processing"""
        try:
            import duckdb
            
            # Try to auto-detect delimiter if not specified
            if delimiter == ',':
                detected_delimiter = self._auto_detect_delimiter(file_path)
                if detected_delimiter:
                    delimiter = detected_delimiter
                    logger.info(f"🔍 Auto-detected delimiter: '{delimiter}'")
            
            # Use DuckDB for direct CSV reading (10-100x faster than Pandas)
            conn = duckdb.connect()
            
            # Read CSV directly into DuckDB (auto-detects types, handles encoding)
            try:
                # DuckDB's read_csv_auto uses 'delim' or 'sep', NOT 'delimiter'
                # Escape single quotes in file path for SQL safety
                safe_file_path = file_path.replace("'", "''")
                conn.execute(f"""
                    CREATE TABLE data AS 
                    SELECT * FROM read_csv_auto('{safe_file_path}', 
                        delim='{delimiter}',
                        header=true,
                        auto_detect=true
                    )
                """)
                
                # Get schema from DuckDB (faster and more accurate)
                schema_result = conn.execute("DESCRIBE data").fetchall()
                schema = {
                    'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                    'row_count': conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                }
                
                # Get sample data (first 100 rows for preview/schema enhancement)
                sample_result = conn.execute("SELECT * FROM data LIMIT 100").fetchall()
                columns = [col[0] for col in schema_result]
                
                # Convert to list of dictionaries
                data = [dict(zip(columns, row)) for row in sample_result]
                
                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)
                
                # Get full row count
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                schema['row_count'] = total_rows
                
                conn.close()
                
                logger.info(f"🦆 Processed CSV with DuckDB: {total_rows} rows, {len(columns)} columns")
                
                return data, schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB CSV read failed, falling back to Pandas: {duckdb_error}")
                conn.close()
                # Fallback to Pandas if DuckDB fails
                df = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding)
                data = df.to_dict('records')
                for row in data:
                    for key, value in row.items():
                        if pd.isna(value):
                            row[key] = None
                
                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)
                
                schema = self._infer_schema_from_dataframe(df)
                return data, schema
            
        except Exception as error:
            raise Exception(f"CSV processing failed: {str(error)}")

    async def _process_parquet_file(self, file_path: str) -> tuple:
        """Process Parquet files using DuckDB for native, fast processing"""
        try:
            import duckdb
            
            # Use DuckDB for direct Parquet reading (native format, fastest)
            conn = duckdb.connect()
            
            try:
                # Read Parquet directly into DuckDB (native format, no conversion needed)
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_parquet('{file_path}')")
                
                # Get schema from DuckDB
                schema_result = conn.execute("DESCRIBE data").fetchall()
                schema = {
                    'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                    'row_count': conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                }
                
                # Get sample data (first 100 rows)
                sample_result = conn.execute("SELECT * FROM data LIMIT 100").fetchall()
                columns = [col[0] for col in schema_result]
                
                # Convert to list of dictionaries
                data = [dict(zip(columns, row)) for row in sample_result]
                
                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)
                
                # Get full row count
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                schema['row_count'] = total_rows
                
                # Try to get Parquet metadata
                try:
                    import pyarrow.parquet as pq
                    parquet_file = pq.ParquetFile(file_path)
                    metadata = parquet_file.metadata
                    schema['parquet_metadata'] = {
                        'row_groups': getattr(metadata, 'num_row_groups', 0),
                        'total_rows': total_rows,
                        'created_by': getattr(metadata, 'created_by', 'unknown'),
                        'schema_version': getattr(metadata, 'schema_version', 'unknown')
                    }
                except Exception:
                    schema['parquet_metadata'] = {
                        'row_groups': 1,
                        'total_rows': total_rows,
                        'created_by': 'unknown',
                        'schema_version': 'unknown'
                    }
                
                conn.close()
                
                logger.info(f"🦆 Processed Parquet with DuckDB: {total_rows} rows, {len(columns)} columns")
                
                return data, schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB Parquet read failed, falling back to PyArrow: {duckdb_error}")
                conn.close()
                # Fallback to PyArrow/Pandas if DuckDB fails
                import pyarrow.parquet as pq
                df = pq.read_table(file_path).to_pandas()
                data = df.to_dict('records')
                for row in data:
                    for key, value in row.items():
                        if pd.isna(value):
                            row[key] = None
                schema = self._infer_schema_from_dataframe(df)
                return data, schema
            
        except Exception as error:
            raise Exception(f"Parquet processing failed: {str(error)}")

    async def _process_excel_file(self, file_path: str, sheet_name: Optional[str] = None) -> tuple:
        """
        Process Excel files with multi-sheet support using DuckDB.
        Creates virtual tables for each sheet, enabling SQL queries across sheets.
        """
        try:
            import duckdb
            import openpyxl  # For reading Excel sheet names
            import re
            
            # Use DuckDB for Excel processing (supports multi-sheet)
            conn = duckdb.connect()
            
            try:
                # Get all sheet names from Excel file
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                except Exception as e:
                    logger.warning(f"Could not read Excel sheet names with openpyxl: {e}, using default")
                    # Fallback: try to read with pandas to get sheet names
                    try:
                        xl_file = pd.ExcelFile(file_path)
                        sheet_names = xl_file.sheet_names
                    except:
                        sheet_names = ['Sheet1']  # Default sheet name
                
                logger.info(f"📊 Found {len(sheet_names)} sheet(s) in Excel file: {sheet_names}")
                
                # Process each sheet and create a virtual table
                all_schemas = {}
                primary_data = None
                primary_schema = None
                
                for idx, sheet in enumerate(sheet_names):
                    try:
                        # Read sheet into pandas DataFrame
                        df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
                        
                        # Clean data (convert NaN to None)
                        df = df.replace({pd.NA: None, pd.NaT: None})
                        df = df.where(pd.notnull(df), None)
                        
                        # Create sanitized table name (DuckDB-safe)
                        table_name = f"sheet_{idx}_{sheet.replace(' ', '_').replace('-', '_').replace('.', '_')[:50]}"
                        # Remove special characters that might break SQL
                        table_name = re.sub(r'[^a-zA-Z0-9_]', '_', table_name)
                        
                        # Register DataFrame in DuckDB
                        conn.register(f"_temp_{table_name}", df)
                        
                        # Create persistent table in DuckDB
                        conn.execute(f"CREATE TABLE {table_name} AS SELECT * FROM _temp_{table_name}")
                        
                        # Get schema for this sheet
                        schema_result = conn.execute(f"DESCRIBE {table_name}").fetchall()
                        sheet_schema = {
                            'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                            'row_count': conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                        }
                        
                        all_schemas[sheet] = {
                            'table_name': table_name,
                            'schema': sheet_schema,
                            'row_count': sheet_schema['row_count']
                        }
                        
                        # Use first sheet as primary data (for backward compatibility)
                        if idx == 0 or (sheet_name and sheet == sheet_name):
                            primary_data = df.to_dict('records')
                            # Convert date/datetime objects to JSON-serializable strings
                            primary_data = self._make_json_serializable(primary_data)
                            primary_schema = self._infer_schema_from_dataframe(df)
                            primary_schema['row_count'] = sheet_schema['row_count']
                            primary_schema['table_name'] = table_name
                            primary_schema['all_sheets'] = all_schemas
                        
                        logger.info(f"✅ Created DuckDB table '{table_name}' for sheet '{sheet}' ({sheet_schema['row_count']} rows)")
                        
                    except Exception as sheet_error:
                        logger.warning(f"⚠️ Failed to process sheet '{sheet}': {sheet_error}")
                        continue
                
                if not primary_data:
                    raise Exception("No sheets could be processed from Excel file")
                
                # Store DuckDB connection info in schema for later use
                primary_schema['duckdb_tables'] = {sheet: info['table_name'] for sheet, info in all_schemas.items()}
                primary_schema['duckdb_connection'] = 'in_memory'  # Mark that tables are in DuckDB
                
                conn.close()
                
                logger.info(f"🦆 Processed Excel with DuckDB: {len(sheet_names)} sheets, primary: {primary_schema['row_count']} rows")
                
                return primary_data, primary_schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB Excel processing failed, falling back to Pandas: {duckdb_error}")
                conn.close()
                # Fallback to original Pandas approach
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(file_path)
                
                data = df.to_dict('records')
                for row in data:
                    for key, value in row.items():
                        if pd.isna(value):
                            row[key] = None
                
                schema = self._infer_schema_from_dataframe(df)
                return data, schema
            
        except Exception as error:
            raise Exception(f"Excel processing failed: {str(error)}")

    async def _process_json_file(self, file_path: str) -> tuple:
        """Process JSON files using DuckDB for fast, direct processing"""
        try:
            import duckdb
            
            # Use DuckDB for direct JSON reading
            conn = duckdb.connect()
            
            try:
                # Read JSON directly into DuckDB (handles nested structures)
                conn.execute(f"CREATE TABLE data AS SELECT * FROM read_json_auto('{file_path}')")
                
                # Get schema from DuckDB
                schema_result = conn.execute("DESCRIBE data").fetchall()
                schema = {
                    'columns': [{'name': col[0], 'type': col[1]} for col in schema_result],
                    'row_count': conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                }
                
                # Get sample data (first 100 rows)
                sample_result = conn.execute("SELECT * FROM data LIMIT 100").fetchall()
                columns = [col[0] for col in schema_result]
                
                # Convert to list of dictionaries
                data = [dict(zip(columns, row)) for row in sample_result]
                
                # Convert date/datetime objects to JSON-serializable strings
                data = self._make_json_serializable(data)
                
                # Get full row count
                total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()[0]
                schema['row_count'] = total_rows
                
                conn.close()
                
                logger.info(f"🦆 Processed JSON with DuckDB: {total_rows} rows, {len(columns)} columns")
                
                return data, schema
                
            except Exception as duckdb_error:
                logger.warning(f"⚠️ DuckDB JSON read failed, falling back to Pandas: {duckdb_error}")
                conn.close()
                # Fallback to Pandas if DuckDB fails
                with open(file_path, 'r', encoding='utf-8') as f:
                    import json
                    json_data = json.load(f)
                
                if isinstance(json_data, list):
                    data = json_data[:100]  # Sample
                    # Convert to DataFrame for schema inference
                    df = pd.DataFrame(data)
                    schema = self._infer_schema_from_dataframe(df)
                else:
                    # Single object - wrap in list
                    data = [json_data]
                    df = pd.DataFrame(data)
                    schema = self._infer_schema_from_dataframe(df)
                
                return data, schema
            
        except Exception as error:
            raise Exception(f"JSON processing failed: {str(error)}")
    
    async def _process_json_file_old(self, file_path: str) -> tuple:
        """Process JSON files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                json_data = json.load(file)
            
            # Handle different JSON structures
            if isinstance(json_data, list):
                data = json_data
            elif isinstance(json_data, dict):
                if 'data' in json_data and isinstance(json_data['data'], list):
                    data = json_data['data']
                else:
                    data = [json_data]
            else:
                raise ValueError('Invalid JSON structure - expected array or object with data array')
            
            # Convert to DataFrame for schema inference
            df = pd.DataFrame(data)
            
            # Infer schema
            schema = self._infer_schema_from_dataframe(df)
            
            return data, schema
            
        except Exception as error:
            raise Exception(f"JSON processing failed: {str(error)}")

    def _infer_schema_from_dataframe(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Infer schema from pandas DataFrame"""
        columns = []
        types = {}
        statistics = {}
        
        for column in df.columns:
            series = df[column]
            
            # Determine data type
            if pd.api.types.is_numeric_dtype(series):
                if pd.api.types.is_integer_dtype(series):
                    data_type = 'integer'
                else:
                    data_type = 'number'
                
                # Calculate statistics for numeric columns
                statistics[column] = {
                    'min': float(series.min()) if not pd.isna(series.min()) else None,
                    'max': float(series.max()) if not pd.isna(series.max()) else None,
                    'mean': float(series.mean()) if not pd.isna(series.mean()) else None,
                    'null_count': int(series.isnull().sum())
                }
            elif pd.api.types.is_datetime64_any_dtype(series):
                data_type = 'date'
                statistics[column] = {
                    'null_count': int(series.isnull().sum())
                }
            elif pd.api.types.is_bool_dtype(series):
                data_type = 'boolean'
                statistics[column] = {
                    'null_count': int(series.isnull().sum())
                }
            else:
                data_type = 'string'
                
                # Calculate statistics for string columns
                non_null_series = series.dropna()
                statistics[column] = {
                    'unique_count': int(series.nunique()),
                    'max_length': int(non_null_series.astype(str).str.len().max()) if len(non_null_series) > 0 else 0,
                    'null_count': int(series.isnull().sum())
                }
            
            types[column] = data_type
            columns.append({
                'name': column,
                'type': data_type,
                'nullable': bool(series.isnull().any()),  # Convert numpy.bool_ to Python bool
                'statistics': statistics[column]
            })
        
        return {
            'columns': columns,
            'types': types,
            'row_count': int(len(df)),  # Convert numpy.int64 to Python int
            'inferred_at': datetime.now().isoformat()
        }

    def _parse_database_uri(self, uri: str) -> Dict[str, Any]:
        """Parse database URI into connection parameters"""
        try:
            from urllib.parse import urlparse
            
            parsed = urlparse(uri)
            
            # Map protocol to database type
            protocol_map = {
                'postgres': 'postgresql',
                'postgresql': 'postgresql',
                'mysql': 'mysql',
                'sqlserver': 'sqlserver',
                'mssql': 'sqlserver'
            }
            
            db_type = protocol_map.get(parsed.scheme, parsed.scheme)
            
            return {
                'type': db_type,
                'host': parsed.hostname,
                'port': parsed.port,
                'database': parsed.path.lstrip('/') if parsed.path else '',
                'username': parsed.username,
                'password': parsed.password,
                'uri': uri
            }
        except Exception as e:
            raise ValueError(f"Invalid database URI format: {str(e)}")

    async def create_database_connection(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Create database connection using Cube.js connectors"""
        try:
            # Handle URI-based connections
            if config.get('uri'):
                logger.info(f"🔌 Parsing database URI connection")
                parsed_config = self._parse_database_uri(config['uri'])
                # Merge with original config, keeping name if provided
                config = {**parsed_config, **{k: v for k, v in config.items() if k != 'uri'}}
                logger.info(f"🔌 Parsed URI to: {config.get('type')} at {config.get('host')}:{config.get('port')}")
            
            logger.info(f"🔌 Creating database connection: {config.get('type')}")
            
            db_type = config.get('type')
            supported_dbs = self.database_connector.get_supported_databases()
            
            if db_type not in supported_dbs:
                raise ValueError(f"Unsupported database type: {db_type}. Supported: {supported_dbs}")
            
            # Create connection using DatabaseConnectorService
            cube_result = await self.database_connector.create_connection(config)
            
            if not cube_result['success']:
                raise Exception(cube_result['error'])
            
            cube_data_source = cube_result['data_source']
            
            # Generate our data source metadata
            data_source = {
                'id': cube_data_source['id'],
                'name': cube_data_source['name'],
                'type': 'database',
                'db_type': cube_data_source['db_type'],
                'host': config.get('host'),
                'database': config.get('database'),
                'created_at': datetime.now().isoformat(),
                'cube_integration': True,
                'driver': cube_data_source['driver'],
                'status': cube_data_source['status']
            }
            
            # Store in registry AND database
            self.data_sources[data_source['id']] = {
                **data_source,
                'cube_data_source': cube_data_source
            }
            
            # Also save to database for persistence
            try:
                from app.db.models import DataSource as DataSourceModel
                from app.db.session import get_db
                
                # Create data source model instance
                db_data_source = DataSourceModel(
                    id=data_source['id'],
                    name=data_source['name'],
                    type='database',
                    db_type=data_source.get('db_type'),
                    connection_config=json.dumps(config),
                    metadata=json.dumps(data_source),
                    is_active=True
                )
                
                # Add to session and commit
                db = get_db()
                await db.add(db_data_source)
                logger.info(f"✅ Data source saved to database: {data_source['id']}")
            except Exception as db_error:
                logger.warning(f"⚠️ Could not save to database: {db_error}")
                # Continue anyway since we have in-memory storage
            
            logger.info(f"✅ Cube.js database connection created successfully: {data_source['id']}")
            
            return {
                'success': True,
                'data_source': {
                    'id': data_source['id'],
                    'name': data_source['name'],
                    'type': data_source['type'],
                    'db_type': data_source['db_type'],
                    'created_at': data_source['created_at'],
                    'cube_integration': True,
                    'driver': data_source['driver'],
                    'status': data_source['status']
                }
            }
            
        except Exception as error:
            logger.error(f"❌ Cube.js database connection failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }

    async def query_data_source(
        self, 
        data_source_id: str, 
        query: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Query data from data source"""
        try:
            data_source = self.data_sources.get(data_source_id)
            if not data_source:
                raise ValueError(f"Data source not found: {data_source_id}")
            
            if query is None:
                query = {}
            
            if data_source['type'] == 'file':
                return await self._query_file_data_source(data_source, query)
            elif data_source['type'] == 'database':
                return await self._query_database_data_source(data_source, query)
            else:
                raise ValueError(f"Unsupported data source type: {data_source['type']}")
                
        except Exception as error:
            logger.error(f"❌ Data source query failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }

    async def _query_file_data_source(
        self, 
        data_source: Dict[str, Any], 
        query: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Query file-based data source"""
        # CRITICAL: Check if data is in memory first
        data = data_source.get('data', [])
        
        # If no in-memory data, try to load from PostgreSQL storage
        if not data or len(data) == 0:
            object_key = data_source.get('file_path')  # Now it's object_key
            if object_key:
                try:
                    from app.modules.data.services.postgres_storage_service import PostgresStorageService
                    storage_service = PostgresStorageService()
                    user_id = data_source.get('user_id')
                    
                    if not user_id:
                        logger.warning("⚠️ user_id not found in data_source, cannot load from PostgreSQL storage")
                    else:
                        # Load file from PostgreSQL
                        logger.info(f"📊 Loading file data from PostgreSQL storage: {object_key}")
                        file_content = await storage_service.get_file(object_key, user_id)
                        
                        # Write to temp file for processing
                        file_format = data_source.get('format', 'csv')
                        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_format}") as tmp:
                            tmp.write(file_content)
                            tmp_path = tmp.name
                        
                        try:
                            # Process file
                            data = await self._read_file_data(tmp_path, file_format, limit=query.get('limit', 10000))
                            logger.info(f"✅ Loaded {len(data)} rows from PostgreSQL storage")
                        finally:
                            # Clean up temp file
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
                except Exception as e:
                    logger.error(f"❌ Failed to load from PostgreSQL storage: {str(e)}")
                    data = []
        
        # Fallback to sample_data
        if not data or len(data) == 0:
            data = data_source.get('sample_data', []) or data_source.get('preview_data', [])
            if data:
                logger.info(f"📊 Using sample/preview data ({len(data)} rows)")
        
        # Apply filters
        filters = query.get('filters', [])
        for filter_item in filters:
            column = filter_item.get('column')
            operator = filter_item.get('operator')
            value = filter_item.get('value')
            
            if operator == 'equals':
                data = [row for row in data if str(row.get(column, '')).lower() == str(value).lower()]
            elif operator == 'contains':
                data = [row for row in data if str(value).lower() in str(row.get(column, '')).lower()]
            elif operator == 'greater_than':
                data = [row for row in data if float(row.get(column, 0) or 0) > float(value)]
            elif operator == 'less_than':
                data = [row for row in data if float(row.get(column, 0) or 0) < float(value)]
        
        # Apply sorting
        sort_config = query.get('sort')
        if sort_config:
            column = sort_config.get('column')
            direction = sort_config.get('direction', 'asc')
            reverse = direction == 'desc'
            
            try:
                data = sorted(data, key=lambda x: x.get(column, ''), reverse=reverse)
            except:
                # If sorting fails, continue without sorting
                pass
        
        # Apply pagination
        offset = query.get('offset', 0)
        limit = query.get('limit', 1000)
        total_rows = len(data)
        paginated_data = data[offset:offset + limit]
        
        return {
            'success': True,
            'data': paginated_data,
            'total_rows': total_rows,
            'offset': offset,
            'limit': limit,
            'schema': data_source.get('schema')
        }

    async def _query_database_data_source(
        self, 
        data_source: Dict[str, Any], 
        query: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Query database data source through Cube.js"""
        try:
            if not data_source.get('cube_integration'):
                return {
                    'success': False,
                    'error': 'Database source not integrated with Cube.js'
                }
            
            cube_data_source = data_source.get('cube_data_source')
            if not cube_data_source:
                return {
                    'success': False,
                    'error': 'Cube.js data source not found'
                }
            
            # Execute query through DatabaseConnectorService
            # Note: query should be a SQL string, not a structured query
            sql_query = query.get('sql') or query.get('query') if isinstance(query, dict) else str(query)
            
            result = await self.database_connector.execute_query(
                data_source['id'],
                sql_query
            )
            
            return result
            
        except Exception as error:
            logger.error(f"❌ Database query failed: {str(error)}")
            return {
                'success': False,
                'error': str(error)
            }

    async def get_data_source(self, data_source_id: str) -> Dict[str, Any]:
        """Get a specific data source by ID - checks both in-memory and database"""
        try:
            # First check in-memory cache
            data_source = self.data_sources.get(data_source_id)
            if data_source:
                logger.info(f"✅ Found data source {data_source_id} in memory cache")
                return {
                    'success': True,
                    'data_source': data_source
                }
            
            # If not in memory, check database
            logger.info(f"🔍 Data source {data_source_id} not in memory, checking database...")
            from app.modules.data.models import DataSource
            from app.db.session import async_session
            import json

            async with async_session() as db:
                from sqlalchemy import select
                
                query = select(DataSource).where(
                    DataSource.id == data_source_id,
                    DataSource.is_active == True
                )
                result = await db.execute(query)
                db_source = result.scalar_one_or_none()
                
                if db_source:
                    logger.info(f"✅ Found data source {data_source_id} in database")
                    # Convert database model to dict format
                    data_source_dict = {
                        'id': str(db_source.id),
                        'name': db_source.name,
                        'type': db_source.type,
                        'format': db_source.format,
                        'description': db_source.description,
                        'user_id': str(db_source.user_id) if db_source.user_id else None,
                        'row_count': db_source.row_count,
                        'size': db_source.size,
                        'is_active': db_source.is_active,
                        'created_at': db_source.created_at.isoformat() if db_source.created_at else None,
                        'updated_at': db_source.updated_at.isoformat() if db_source.updated_at else None,
                    }
                    
                    # Add file_path if available directly from model (now stores object_key)
                    if hasattr(db_source, 'file_path') and db_source.file_path:
                        data_source_dict['file_path'] = db_source.file_path  # This is now object_key
                    
                    # Parse config and schema if available
                    try:
                        if db_source.connection_config:
                            config = json.loads(db_source.connection_config) if isinstance(db_source.connection_config, str) else db_source.connection_config
                            data_source_dict['config'] = config
                            data_source_dict['connection_config'] = config
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to parse connection_config: {e}")
                    
                    try:
                        if db_source.schema:
                            schema = json.loads(db_source.schema) if isinstance(db_source.schema, str) else db_source.schema
                            data_source_dict['schema'] = schema
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to parse schema: {e}")
                    
                    # Load sample_data from database
                    if db_source.sample_data:
                        try:
                            sample = json.loads(db_source.sample_data) if isinstance(db_source.sample_data, str) else db_source.sample_data
                            data_source_dict['sample_data'] = sample
                            data_source_dict['data'] = sample  # For compatibility
                            logger.info(f"✅ Loaded sample_data from database")
                        except Exception as e:
                            logger.warning(f"⚠️ Failed to parse sample_data: {e}")
                    
                    # Add original_filename if available
                    if hasattr(db_source, 'original_filename') and db_source.original_filename:
                        data_source_dict['original_filename'] = db_source.original_filename
                    
                    # Cache in memory for future access
                    self.data_sources[data_source_id] = data_source_dict
                    
                    return {
                        'success': True,
                        'data_source': data_source_dict
                    }
                else:
                    logger.warning(f"⚠️ Data source {data_source_id} not found in database or memory")
                    return {
                        'success': False,
                        'error': 'Data source not found'
                    }
        except Exception as e:
            logger.error(f"❌ Failed to get data source: {str(e)}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }

    def list_data_sources(self) -> Dict[str, Any]:
        """List all data sources"""
        sources = []
        for source in self.data_sources.values():
            metadata = {key: value for key, value in source.items() 
                       if key not in ['data', 'connection']}
            sources.append(metadata)
        
        return {
            'success': True,
            'data_sources': sources,
            'count': len(sources)
        }

    def delete_data_source(self, data_source_id: str) -> Dict[str, Any]:
        """Delete data source"""
        try:
            data_source = self.data_sources.get(data_source_id)
            
            # Attempt DB deletion first (if persisted)
            try:
                from app.db.models import DataSource
                from app.modules.projects.models import ProjectDataSource
                # Use async_session factory directly (get_async_session is an async generator for FastAPI deps)
                from app.db.session import async_session
                import asyncio
                async def _delete_from_db():
                    async with async_session() as db:
                        from sqlalchemy import delete
                        # Remove project links
                        try:
                            await db.execute(delete(ProjectDataSource).where(ProjectDataSource.data_source_id == data_source_id))
                        except Exception:
                            pass
                        # Remove data source row
                        await db.execute(delete(DataSource).where(DataSource.id == data_source_id))
                        await db.commit()
                # Run async deletion in current loop if available
                try:
                    loop = asyncio.get_event_loop()
                    if loop.is_running():
                        fut = asyncio.ensure_future(_delete_from_db())
                        # fire-and-forget
                    else:
                        loop.run_until_complete(_delete_from_db())
                except RuntimeError:
                    # No loop, run new loop
                    asyncio.run(_delete_from_db())
            except Exception as db_del_err:
                logger.warning(f"DB deletion for data source {data_source_id} skipped/failed: {db_del_err}")

            # In-memory cleanup
            if data_source:
                # Clean up file if it's a file-based source
                if data_source.get('type') == 'file' and 'file_path' in data_source:
                    file_path = data_source['file_path']
                    if os.path.exists(file_path):
                        os.unlink(file_path)
                # Close database connection if tracked
                if data_source.get('type') == 'database' and 'connection' in data_source:
                    connection = data_source['connection']
                    if hasattr(connection, 'close'):
                        try:
                            connection.close()
                        except Exception:
                            pass
                self.data_sources.pop(data_source_id, None)

            return {'success': True, 'message': 'Data source deleted successfully'}
        except Exception as e:
            logger.error(f"Failed to delete data source {data_source_id}: {e}")
            return {'success': False, 'error': str(e)}

    async def generate_data_insights(self, data_source_id: str) -> Dict[str, Any]:
        """Generate AI insights for a data source"""
        try:
            logger.info(f"🔍 Generating AI insights for data source: {data_source_id}")
            
            # Get the data source
            data_source = self.data_sources.get(data_source_id)
            if not data_source:
                return {'success': False, 'error': 'Data source not found'}
            
            # Get data and schema
            data = data_source.get('data', [])
            schema = data_source.get('schema', {})
            name = data_source.get('name', 'Unknown')
            
            # Generate insights using AI
            insights_result = await self.ai_schema_service.generate_data_insights(
                data, schema, name
            )
            
            return insights_result
            
        except Exception as e:
            logger.error(f"❌ Failed to generate insights: {str(e)}")
            return {'success': False, 'error': str(e)}

    async def get_database_schema(self, data_source_id: str) -> Dict[str, Any]:
        """Get database schema information for a connected database"""
        try:
            logger.info(f"🔍 Fetching database schema for: {data_source_id}")
            
            # Get the data source from database
            from app.modules.data.models import DataSource
            from app.db.session import async_session

            async with async_session() as db:
                from sqlalchemy import select
                
                query = select(DataSource).where(DataSource.id == data_source_id)
                result = await db.execute(query)
                data_source = result.scalar_one_or_none()
                
                if not data_source:
                    return {
                        'success': False,
                        'error': 'Data source not found'
                    }
                
                if data_source.type != 'database' and data_source.type != 'warehouse':
                    return {
                        'success': False,
                        'error': 'Data source is not a database or warehouse connection'
                    }
                
                # Parse the stored schema/config
                try:
                    config = json.loads(data_source.connection_config) if data_source.connection_config else {}
                    schema_info = json.loads(data_source.schema) if data_source.schema else {}
                except json.JSONDecodeError:
                    config = {}
                    schema_info = {}
                
                # CRITICAL: Decrypt credentials before using for schema fetch
                try:
                    from app.modules.data.utils.credentials import decrypt_credentials
                    config = decrypt_credentials(config)
                    logger.debug(f"✅ Decrypted credentials for schema fetch: {data_source_id}")
                except Exception as decrypt_error:
                    logger.warning(f"⚠️ Could not decrypt credentials for schema fetch (may not be encrypted): {decrypt_error}")
                
                # Try to get live schema from the database
                try:
                    logger.info(f"🔍 Fetching live schema for {data_source_id}, db_type: {data_source.db_type}, type: {data_source.type}")
                    live_schema = await self._fetch_live_database_schema(config)
                    logger.info(f"📊 Live schema result: success={live_schema.get('success')}, tables_count={len(live_schema.get('tables', []))}, schemas_count={len(live_schema.get('schemas', []))}")
                    
                    if live_schema['success']:
                        tables = live_schema.get('tables', [])
                        schemas = live_schema.get('schemas', [])
                        
                        # Update the stored schema with live data
                        updated_schema = {
                            **schema_info,
                            'tables': tables,
                            'schemas': schemas,
                            'last_updated': datetime.now().isoformat()
                        }
                        
                        logger.info(f"✅ Schema fetched successfully: {len(tables)} tables, {len(schemas)} schemas")
                        
                        # Update the database record
                        data_source.schema = json.dumps(updated_schema)
                        data_source.row_count = live_schema.get('total_rows', 0)
                        data_source.updated_at = datetime.now()
                        await db.commit()
                        
                        return {
                            'success': True,
                            'schema': updated_schema,
                            'data_source': {
                                'id': data_source.id,
                                'name': data_source.name,
                                'type': data_source.type,
                                'db_type': data_source.db_type,
                                'row_count': data_source.row_count
                            }
                        }
                    else:
                        # Return stored schema if live fetch fails
                        return {
                            'success': True,
                            'schema': schema_info,
                            'data_source': {
                                'id': data_source.id,
                                'name': data_source.name,
                                'type': data_source.type,
                                'db_type': data_source.db_type,
                                'row_count': data_source.row_count
                            },
                            'warning': 'Using cached schema - live fetch failed'
                        }
                        
                except Exception as live_error:
                    logger.warning(f"⚠️ Live schema fetch failed: {str(live_error)}")
                    # Return stored schema
                    return {
                        'success': True,
                        'schema': schema_info,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'db_type': data_source.db_type,
                            'row_count': data_source.row_count
                        },
                        'warning': 'Using cached schema - live fetch failed'
                    }
                    
        except Exception as e:
            logger.error(f"❌ Failed to get database schema: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    async def _fetch_live_database_schema(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Fetch live schema from the database connection"""
        try:
            # CRITICAL: Ensure credentials are decrypted before using
            try:
                from app.modules.data.utils.credentials import decrypt_credentials
                config = decrypt_credentials(config)
                logger.debug(f"✅ Decrypted credentials for schema fetch")
            except Exception as decrypt_error:
                logger.debug(f"Credentials may not be encrypted: {decrypt_error}")
            
            db_type = config.get('type', '').lower()
            
            # Get supported databases from DatabaseConnectorService
            supported_dbs = self.database_connector.get_supported_databases()
            
            if db_type not in supported_dbs:
                return {
                    'success': False,
                    'error': f'Unsupported database type: {db_type}. Supported: {supported_dbs}'
                }
            
            # Use DatabaseConnectorService to get schema
            try:
                schema_result = await self.database_connector.get_schema(config)
                if schema_result['success']:
                    return {
                        'success': True,
                        'tables': schema_result.get('tables', []),
                        'schemas': schema_result.get('schemas', []),
                        'total_rows': schema_result.get('total_rows', 0)
                    }
                else:
                    logger.warning(f"⚠️ Schema fetch failed: {schema_result.get('error')}")
                    # Enhanced fallback with more realistic schema structure
                    return await self._get_enhanced_fallback_schema(config)
                    
            except Exception as schema_error:
                logger.warning(f"⚠️ Schema fetch failed: {str(schema_error)}")
                # Enhanced fallback with more realistic schema structure
                return await self._get_enhanced_fallback_schema(config)
                
        except Exception as e:
            logger.error(f"❌ Live schema fetch failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def _get_enhanced_fallback_schema(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """Get fallback schema - only returns empty schema, no mock data"""
        try:
            db_type = config.get('type', '').lower()
            db_name = config.get('database', 'unknown')
            
            # CRITICAL: Do not return mock data - only return empty schema
            # Real schema should be fetched via get_database_schema which uses SQLAlchemy
            logger.warning(f"⚠️ Fallback schema called for {db_type} - returning empty schema. Real schema should be fetched via database connection.")
            
            # Return empty schema - let the real connection path handle schema fetching
            return {
                'success': True,
                'tables': [],
                'schemas': [],
                'total_rows': 0,
                'warning': 'Schema could not be fetched. Please ensure database connection is valid and try again.'
            }
            
        except Exception as e:
            logger.error(f"❌ Enhanced fallback schema failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    # Utility methods
    def _get_file_extension(self, filename: str) -> str:
        """Get file extension"""
        return Path(filename).suffix.lower().lstrip('.')

    # Database connector methods
    async def _create_database_connector(self, config: Dict[str, Any]):
        """Create database connector"""
        # This is handled by the database_connector service
        return await self.database_connector.create_connection(config)

    def _convert_to_cube_query(self, query: Dict[str, Any], cube_data_source: Dict[str, Any]) -> Dict[str, Any]:
        """Convert our query format to Cube.js query format"""
        cube_query = {}
        
        # Add measures (default to count if none specified)
        cube_query['measures'] = query.get('measures', ['count'])
        
        # Add dimensions
        if query.get('dimensions'):
            cube_query['dimensions'] = query['dimensions']
        
        # Add time dimensions
        if query.get('time_dimensions'):
            cube_query['timeDimensions'] = query['time_dimensions']
        
        # Add filters
        if query.get('filters'):
            cube_query['filters'] = self._convert_filters_to_cube_format(query['filters'])
        
        # Add sorting
        if query.get('sort'):
            sort_config = query['sort']
            cube_query['order'] = [[sort_config['column'], sort_config.get('direction', 'asc')]]
        
        # Add limit
        if query.get('limit'):
            cube_query['limit'] = query['limit']
        
        return cube_query

    def _convert_filters_to_cube_format(self, filters: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Convert filters to Cube.js format"""
        cube_filters = []
        
        for filter_item in filters:
            cube_filter = {
                'member': filter_item['column'],
                'operator': self._map_operator_to_cube(filter_item['operator']),
                'values': [filter_item['value']]
            }
            cube_filters.append(cube_filter)
        
        return cube_filters

    def _map_operator_to_cube(self, operator: str) -> str:
        """Map our operators to Cube.js operators"""
        operator_mapping = {
            'equals': 'equals',
            'contains': 'contains',
            'greater_than': 'gt',
            'less_than': 'lt',
            'greater_equal': 'gte',
            'less_equal': 'lte'
        }
        return operator_mapping.get(operator, 'equals')

    async def get_supported_databases(self) -> Dict[str, Any]:
        """Get supported database types"""
        supported_dbs = self.database_connector.get_supported_databases()
        return {
            'success': True,
            'supported_databases': [
                {
                    'type': db_type,
                    'name': db_type.title(),
                    'driver': 'sqlalchemy',
                }
                for db_type in supported_dbs
            ]
        }

    def _auto_detect_delimiter(self, file_path: str) -> str:
        """Auto-detect CSV delimiter"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline().strip()
            
            if not first_line:
                return ','
            
            # Common delimiters to test
            delimiters = [',', ';', '\t', '|', ' ']
            max_fields = 0
            best_delimiter = ','
            
            for delimiter in delimiters:
                fields = first_line.split(delimiter)
                if len(fields) > max_fields:
                    max_fields = len(fields)
                    best_delimiter = delimiter
            
            return best_delimiter
            
        except Exception:
            return ','

    def _auto_detect_encoding(self, file_path: str) -> str:
        """Auto-detect file encoding"""
        try:
            import chardet
            
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Read first 10KB
            
            result = chardet.detect(raw_data)
            encoding = result['encoding']
            confidence = result['confidence']
            
            if confidence > 0.7:
                return encoding
            else:
                return 'utf-8'
                
        except ImportError:
            # chardet not available, try common encodings
            encodings = ['utf-8', 'latin-1', 'cp1252']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        f.readline()
                    return encoding
                except UnicodeDecodeError:
                    continue
            
            return 'utf-8'
        except Exception:
            return 'utf-8'

    async def get_file_preview(self, file_path: str, file_format: str, options: dict = None) -> dict:
        """Get enhanced file preview with auto-detection"""
        try:
            if options is None:
                options = {}
            
            preview_data = {
                'format': file_format,
                'size': os.path.getsize(file_path),
                'auto_detected_config': {},
                'sample_data': [],
                'schema_preview': {},
                'processing_options': {}
            }
            
            if file_format in ['csv', 'tsv']:
                # Auto-detect delimiter and encoding
                delimiter = options.get('delimiter', self._auto_detect_delimiter(file_path))
                encoding = options.get('encoding', self._auto_detect_encoding(file_path))
                
                preview_data['auto_detected_config'] = {
                    'delimiter': delimiter,
                    'encoding': encoding
                }
                
                # Read first few lines for preview
                df_preview = pd.read_csv(file_path, delimiter=delimiter, encoding=encoding, nrows=10)
                preview_data['sample_data'] = df_preview.to_dict('records')
                preview_data['schema_preview'] = {
                    'columns': list(df_preview.columns),
                    'data_types': df_preview.dtypes.to_dict(),
                    'row_count_preview': len(df_preview)
                }
                
                # Get processing options
                preview_data['processing_options'] = {
                    'delimiters': self.file_processing_configs['csv']['delimiters'],
                    'encodings': self.file_processing_configs['csv']['encodings']
                }
                
            elif file_format in ['xlsx', 'xls']:
                # Get sheet information
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                sheets = workbook.sheetnames
                
                preview_data['auto_detected_config'] = {
                    'sheets': sheets,
                    'default_sheet': sheets[0] if sheets else None
                }
                
                # Read first sheet for preview
                df_preview = pd.read_excel(file_path, sheet_name=sheets[0], nrows=10)
                preview_data['sample_data'] = df_preview.to_dict('records')
                preview_data['schema_preview'] = {
                    'columns': list(df_preview.columns),
                    'data_types': df_preview.dtypes.to_dict(),
                    'row_count_preview': len(df_preview)
                }
                
                preview_data['processing_options'] = {
                    'sheets': sheets,
                    'header_row_options': [0, 1, 2],
                    'skip_rows_options': [0, 1, 2, 3]
                }
                
            elif file_format == 'parquet':
                # Get parquet metadata
                import pyarrow.parquet as pq
                parquet_file = pq.ParquetFile(file_path)
                metadata = parquet_file.metadata
                
                preview_data['auto_detected_config'] = {
                    'columns': list(metadata.schema.names),
                    'row_groups': metadata.num_row_groups,
                    'total_rows': metadata.num_rows
                }
                
                # Read sample data
                df_preview = parquet_file.read().to_pandas().head(10)
                preview_data['sample_data'] = df_preview.to_dict('records')
                preview_data['schema_preview'] = {
                    'columns': list(df_preview.columns),
                    'data_types': df_preview.dtypes.to_dict(),
                    'row_count_preview': len(df_preview)
                }
                
            elif file_format == 'json':
                # Get JSON structure info
                with open(file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                
                if isinstance(json_data, list):
                    preview_data['auto_detected_config'] = {
                        'structure': 'array',
                        'item_count': len(json_data)
                    }
                    sample_data = json_data[:10] if len(json_data) > 10 else json_data
                else:
                    preview_data['auto_detected_config'] = {
                        'structure': 'object',
                        'keys': list(json_data.keys())
                    }
                    sample_data = [json_data]
                
                preview_data['sample_data'] = sample_data
                preview_data['schema_preview'] = {
                    'structure_type': preview_data['auto_detected_config']['structure'],
                    'sample_keys': list(sample_data[0].keys()) if sample_data else []
                }
            
            return {
                'success': True,
                'preview': preview_data
            }
            
        except Exception as e:
            logger.error(f"❌ File preview generation failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _initialize_demo_data(self):
        """Initialize demo data sources for testing and demonstration"""
        try:
            logger.info("🚀 Initializing demo data sources...")
            
            # Demo sales data
            demo_sales = {
                'id': 'demo_sales_data',
                'name': 'Demo Sales Data',
                'type': 'file',
                'format': 'csv',
                'description': 'Sample sales data for demonstration and testing',
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'size': 24576,  # 24KB
                'row_count': 1000,
                'schema': {
                    'columns': [
                        {'name': 'date', 'type': 'datetime', 'nullable': False},
                        {'name': 'product_name', 'type': 'string', 'nullable': False},
                        {'name': 'category', 'type': 'string', 'nullable': False},
                        {'name': 'sales_amount', 'type': 'numeric', 'nullable': False},
                        {'name': 'quantity', 'type': 'integer', 'nullable': False},
                        {'name': 'region', 'type': 'string', 'nullable': False},
                        {'name': 'customer_id', 'type': 'string', 'nullable': False}
                    ]
                },
                'sample_data': [
                    {'date': '2024-01-15', 'product_name': 'Laptop Pro', 'category': 'Electronics', 'sales_amount': 1299.99, 'quantity': 1, 'region': 'North', 'customer_id': 'C001'},
                    {'date': '2024-01-15', 'product_name': 'Wireless Mouse', 'category': 'Accessories', 'sales_amount': 29.99, 'quantity': 2, 'region': 'South', 'customer_id': 'C002'},
                    {'date': '2024-01-16', 'product_name': 'Monitor 4K', 'category': 'Electronics', 'sales_amount': 599.99, 'quantity': 1, 'region': 'East', 'customer_id': 'C003'},
                    {'date': '2024-01-16', 'product_name': 'Keyboard', 'category': 'Accessories', 'sales_amount': 89.99, 'quantity': 1, 'region': 'West', 'customer_id': 'C004'},
                    {'date': '2024-01-17', 'product_name': 'Tablet', 'category': 'Electronics', 'sales_amount': 399.99, 'quantity': 1, 'region': 'North', 'customer_id': 'C005'}
                ],
                'metadata': {
                    'source': 'demo_data',
                    'business_domain': 'retail',
                    'data_quality': 'high',
                    'last_updated': datetime.now().isoformat()
                }
            }
            
            # Demo customer data
            demo_customers = {
                'id': 'demo_customers_data',
                'name': 'Demo Customer Data',
                'type': 'file',
                'format': 'csv',
                'description': 'Sample customer data for demonstration and testing',
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat(),
                'size': 15360,  # 15KB
                'row_count': 500,
                'schema': {
                    'columns': [
                        {'name': 'customer_id', 'type': 'string', 'nullable': False},
                        {'name': 'first_name', 'type': 'string', 'nullable': False},
                        {'name': 'last_name', 'type': 'string', 'nullable': False},
                        {'name': 'email', 'type': 'string', 'nullable': False},
                        {'name': 'age', 'type': 'integer', 'nullable': True},
                        {'name': 'city', 'type': 'string', 'nullable': False},
                        {'name': 'country', 'type': 'string', 'nullable': False},
                        {'name': 'registration_date', 'type': 'datetime', 'nullable': False}
                    ]
                },
                'sample_data': [
                    {'customer_id': 'C001', 'first_name': 'John', 'last_name': 'Doe', 'email': 'john.doe@email.com', 'age': 35, 'city': 'New York', 'country': 'USA', 'registration_date': '2023-01-15'},
                    {'customer_id': 'C002', 'first_name': 'Jane', 'last_name': 'Smith', 'email': 'jane.smith@email.com', 'age': 28, 'city': 'Los Angeles', 'country': 'USA', 'registration_date': '2023-02-20'},
                    {'customer_id': 'C003', 'first_name': 'Bob', 'last_name': 'Johnson', 'email': 'bob.johnson@email.com', 'age': 42, 'city': 'Chicago', 'country': 'USA', 'registration_date': '2023-03-10'},
                    {'customer_id': 'C004', 'first_name': 'Alice', 'last_name': 'Brown', 'email': 'alice.brown@email.com', 'age': 31, 'city': 'Houston', 'country': 'USA', 'registration_date': '2023-04-05'},
                    {'customer_id': 'C005', 'first_name': 'Charlie', 'last_name': 'Wilson', 'email': 'charlie.wilson@email.com', 'age': 39, 'city': 'Phoenix', 'country': 'USA', 'registration_date': '2023-05-12'}
                ],
                'metadata': {
                    'source': 'demo_data',
                    'business_domain': 'customer_management',
                    'data_quality': 'high',
                    'last_updated': datetime.now().isoformat()
                }
            }
            
            # Add demo data sources to the registry
            self.data_sources['demo_sales_data'] = demo_sales
            self.data_sources['demo_customers_data'] = demo_customers

            # Also expose convenient aliases commonly used by the UI
            self.data_sources['duckdb_local'] = self.data_sources['demo_sales_data']
            self.data_sources['csv_sales'] = self.data_sources['demo_sales_data']
            self.data_sources['snowflake_warehouse'] = self.data_sources['demo_customers_data']
            self.data_sources['postgresql_prod'] = self.data_sources['demo_customers_data']
            
            logger.info(f"✅ Demo data sources initialized successfully: {list(self.data_sources.keys())}")
            logger.info(f"✅ Total demo sources: {len(self.data_sources)}")
            
        except Exception as e:
            logger.error(f"❌ Failed to initialize demo data: {str(e)}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
    
    def _get_current_user_id(self) -> int:
        """Get current user ID from auth context"""
        try:
            # TODO: Implement proper auth context extraction
            # For now, return 1 as default user
            # In production, this should extract user_id from JWT token or request context
            return 1
        except Exception as e:
            logger.warning(f"Failed to get current user ID: {e}, using default")
            return 1

    # 🏗️ PROJECT-SCOPED DATA SOURCE METHODS

    async def get_project_data_sources(
        self, 
        organization_id: str, 
        project_id: str, 
        user_id: str = None, 
        offset: int = 0, 
        limit: int = 100
    ) -> List[Dict[str, Any]]:
        """Get data sources for a specific project (project-scoped)"""
        try:
            logger.info(f"🔍 Getting project data sources for project {project_id} in organization {organization_id}")
            
            from app.modules.data.models import DataSource
            from app.modules.projects.models import ProjectDataSource
            from app.db.session import async_session

            async with async_session() as db:
                from sqlalchemy import select, join

                # Join data sources with project data sources
                query = (
                    select(DataSource)
                    .join(ProjectDataSource, DataSource.id == ProjectDataSource.data_source_id)
                    .where(
                        # Accept numeric IDs or slugs; if not int, skip casting
                        ProjectDataSource.project_id == (int(project_id) if str(project_id).isdigit() else ProjectDataSource.project_id),
                        DataSource.is_active == True,
                        ProjectDataSource.is_active == True
                    )
                    .offset(offset)
                    .limit(limit)
                )
                
                result = await db.execute(query)
                data_sources = result.scalars().all()
                
                return [
                    {
                        'id': ds.id,
                        'name': ds.name,
                        'type': ds.type,
                        'format': ds.format,
                        'size': ds.size,
                        'row_count': ds.row_count,
                        'schema': ds.schema,
                        'organization_id': organization_id,
                        'project_id': project_id,
                        'created_at': ds.created_at.isoformat() if ds.created_at else None,
                        'updated_at': ds.updated_at.isoformat() if ds.updated_at else None,
                        'last_accessed': ds.last_accessed.isoformat() if ds.last_accessed else None
                    }
                    for ds in data_sources
                ]
        except Exception as e:
            logger.error(f"❌ Failed to get project data sources: {str(e)}")
            return []

    async def create_project_data_source(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_data: Dict[str, Any],
        user_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Create a new data source for a specific project
        
        NOTE: This method is legacy and may be deprecated. user_id should be provided.
        """
        # Extract user_id from data_source_data if not provided directly
        if not user_id:
            user_id = data_source_data.get('user_id')
        
        if not user_id:
            raise ValueError("user_id is required for data source creation. Provide it in data_source_data or as a parameter.")
        
        try:
            logger.info(f"📊 Creating project data source for project {project_id} (user_id={user_id})")
            
            from app.modules.data.models import DataSource
            from app.modules.projects.models import ProjectDataSource
            from app.db.session import async_session

            # Generate a stable unique id for this create call so we can
            # expose a preview entry immediately in the in-memory registry
            import uuid
            generated_id = f"ds_{uuid.uuid4().hex}"

            # Store a preview entry so immediate calls (modeling) can read sample data
            try:
                self.data_sources[generated_id] = {
                    'id': generated_id,
                    'name': data_source_data.get('name'),
                    'type': data_source_data.get('type'),
                    'format': data_source_data.get('format'),
                    'schema': data_source_data.get('schema'),
                    'data': data_source_data.get('data') or data_source_data.get('sample_data') or [],
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat(),
                    'organization_id': organization_id,
                    'project_id': project_id,
                    'is_active': True,
                }
            except Exception:
                pass

            async def _do_create():
                async with async_session() as db:
                    # Create the data source
                    # Build metadata and connection config safely
                    metadata = data_source_data.get('metadata', {}) or {}
                    # carry over description into metadata to avoid unknown kwargs on model
                    if data_source_data.get('description'):
                        metadata.setdefault('description', data_source_data.get('description'))

                    connection_config = data_source_data.get('config') or {}
                    # Decrypt any encrypted credentials before using
                    try:
                        from app.modules.data.utils.credentials import decrypt_credentials

                        connection_config = decrypt_credentials(connection_config)
                    except Exception:
                        pass
                    db_type = connection_config.get('type') or data_source_data.get('db_type')

                    # Prepare data source values so we can use them in both
                    # async and sync (test) paths.
                    data_source_values = {
                        'id': generated_id,
                        'name': data_source_data['name'],
                        'type': data_source_data['type'],
                        'format': data_source_data.get('format'),
                        'db_type': db_type,
                        'schema': data_source_data.get('schema'),
                        'connection_config': connection_config,
                    # Allow callers to provide inline sample rows for file sources
                        'sample_data': data_source_data.get('data') or data_source_data.get('sample_data') or [],
                        # Use actual user_id (required for data isolation)
                        'user_id': str(user_id),
                        'is_active': True,
                        'created_at': datetime.now(),
                        'updated_at': datetime.now(),
                    }

                    # create SQLAlchemy DataSource instance for async path
                    # If the generated id already exists due to previous attempts, reuse it
                    try:
                        from app.db.session import get_sync_engine
                        eng_check = get_sync_engine()
                        import sqlalchemy as sa
                        with eng_check.connect() as conn:
                            row = conn.execute(sa.text("SELECT id FROM data_sources WHERE id = :id"), {"id": generated_id}).fetchone()
                            if row and row[0]:
                                # already exists; return early using existing id
                                return {
                                    'success': True,
                                    'data_source': {
                                        'id': str(row[0]),
                                        'name': data_source_values.get('name'),
                                        'type': data_source_values.get('type'),
                                        'organization_id': organization_id,
                                        'project_id': project_id,
                                    },
                                    'data_source_id': str(row[0]),
                                    'id': str(row[0]),
                                    'message': 'Data source already exists'
                                }
                    except Exception:
                        # non-fatal; continue to create
                        pass

                    data_source = DataSource(**data_source_values)

                    # Persist using the global database helper which creates its
                    # own session per call and serializes operations.
                    from app.db.session import get_db
                    from sqlalchemy.exc import IntegrityError as SAIntegrityError

                    db_helper = get_db()
                    # Save data_source via helper (creates new session internally)
                    # Retry on unique key collisions by regenerating id
                    max_attempts = 3
                    attempt = 0
                    while True:
                        try:
                            attempt += 1
                            await db_helper.add(data_source)
                            break
                        except SAIntegrityError as sie:
                            if attempt >= max_attempts:
                                raise
                            # regenerate id and retry
                            import uuid as _uuid
                            new_id = f"ds_{_uuid.uuid4().hex}"
                            data_source_values['id'] = new_id
                            # recreate SQLAlchemy instance with new id
                            data_source = DataSource(**data_source_values)
                            continue
                    # Persist sample_data to the DB if provided
                    try:
                        if data_source_values.get('sample_data'):
                            upd = sa.text("UPDATE data_sources SET sample_data = :sd WHERE id = :id")
                            with get_sync_engine().begin() as sync_conn:
                                sync_conn.execute(upd, {"sd": json.dumps(data_source_values.get('sample_data')), "id": data_source.id})
                    except Exception:
                        # Non-fatal; sample persistence best-effort
                        logger.debug('Failed to persist sample_data for %s', data_source.id)

                    # Link to project and persist
                    project_data_source = ProjectDataSource(
                        project_id=int(project_id),
                        data_source_id=data_source.id,
                        data_source_type=data_source_data['type'],
                        is_active=True,
                        added_at=datetime.now(),
                    )

                    await db_helper.add(project_data_source)

                    return {
                        'success': True,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'organization_id': organization_id,
                            'project_id': project_id,
                        },
                        'data_source_id': data_source.id,
                        'id': data_source.id,
                        'message': 'Data source created successfully',
                    }

            # Use global async write queue to serialize DB writes safely.
            # Execute the create via a sync worker (thread) using the sync engine
            # to avoid asyncpg 'another operation is in progress' errors when
            # other async operations are running on the same event loop.
            try:
                from app.db.write_queue import write_queue

                # Prepare a sync callable that mirrors the async create path
                def _do_create_sync():
                    try:
                        # Rebuild connection/config values
                        connection_config = data_source_data.get('config') or {}
                        try:
                            from app.modules.data.utils.credentials import decrypt_credentials

                            connection_config = decrypt_credentials(connection_config)
                        except Exception:
                            pass

                        db_type = connection_config.get('type') or data_source_data.get('db_type')
                        ds_vals = {
                            'id': generated_id,
                            'name': data_source_data['name'],
                            'type': data_source_data['type'],
                            'format': data_source_data.get('format'),
                            'db_type': db_type,
                            'schema': data_source_data.get('schema'),
                            'connection_config': connection_config,
                            'sample_data': data_source_data.get('data') or data_source_data.get('sample_data') or [],
                            'user_id': str(organization_id),
                            'is_active': True,
                            'created_at': datetime.now(),
                            'updated_at': datetime.now()
                        }

                        from app.db.session import get_sync_engine
                        eng = get_sync_engine()
                        with eng.begin() as conn:
                            insert_ds = sa.text(
                                "INSERT INTO data_sources (id, name, type, format, db_type, size, row_count, schema, description, connection_config, file_path, original_filename, created_at, updated_at, user_id, is_active, last_accessed) "
                                "VALUES (:id, :name, :type, :format, :db_type, :size, :row_count, :schema, :description, :connection_config, :file_path, :original_filename, :created_at, :updated_at, :user_id, :is_active, :last_accessed) "
                                "ON CONFLICT (id) DO NOTHING"
                            )
                            params = {
                                'id': ds_vals.get('id'),
                                'name': ds_vals.get('name'),
                                'type': ds_vals.get('type'),
                                'format': ds_vals.get('format'),
                                'db_type': ds_vals.get('db_type'),
                                'size': ds_vals.get('size'),
                                'row_count': ds_vals.get('row_count'),
                                'schema': json.dumps(ds_vals.get('schema')) if ds_vals.get('schema') is not None else None,
                                'description': ds_vals.get('description') if ds_vals.get('description') is not None else None,
                                'connection_config': json.dumps(ds_vals.get('connection_config')) if ds_vals.get('connection_config') is not None else None,
                                'file_path': ds_vals.get('file_path') if ds_vals.get('file_path') is not None else None,
                                'original_filename': ds_vals.get('original_filename') if ds_vals.get('original_filename') is not None else None,
                                'created_at': ds_vals.get('created_at'),
                                'updated_at': ds_vals.get('updated_at'),
                                'user_id': str(ds_vals.get('user_id')) if ds_vals.get('user_id') is not None else None,
                                'is_active': ds_vals.get('is_active'),
                                'last_accessed': ds_vals.get('last_accessed')
                            }
                            # Retry insert on unique violation
                            max_sync_attempts = 3
                            sync_attempt = 0
                            while True:
                                try:
                                    conn.execute(insert_ds, params)

                                    # ensure id exists (select back)
                                    sel = sa.text("SELECT id FROM data_sources WHERE id = :id LIMIT 1")
                                    rsel = conn.execute(sel, {"id": params['id']})
                                    row = rsel.fetchone()
                                    if not row:
                                        # nothing inserted (conflict); try to find by name+project
                                        sel2 = sa.text(
                                            "SELECT ds.id FROM data_sources ds JOIN project_data_source pds ON pds.data_source_id = ds.id "
                                            "WHERE ds.name = :name AND pds.project_id = :pid LIMIT 1"
                                        )
                                        r2 = conn.execute(sel2, {"name": ds_vals.get('name'), "pid": int(project_id)})
                                        rr = r2.fetchone()
                                        if rr and rr[0]:
                                            params['id'] = rr[0]
                                            break
                                    else:
                                        # inserted successfully
                                        break
                                except Exception as e:
                                    # detect unique violation (asyncpg / psycopg2 messages vary)
                                    msg = str(e).lower()
                                    if 'unique' in msg or 'duplicate' in msg:
                                        sync_attempt += 1
                                        if sync_attempt >= max_sync_attempts:
                                            raise
                                        import uuid as _uuid
                                        new_id = f"ds_{_uuid.uuid4().hex}"
                                        params['id'] = new_id
                                        ds_vals['id'] = new_id
                                        # loop and retry with new id
                                        continue
                                    raise
                            # Persist sample_data if present
                            if ds_vals.get('sample_data'):
                                try:
                                    upd = sa.text("UPDATE data_sources SET sample_data = :sd WHERE id = :id")
                                    conn.execute(upd, {"sd": json.dumps(ds_vals.get('sample_data')), "id": ds_vals.get('id')})
                                except Exception:
                                    logger.debug('Failed to persist sample_data in sync path for %s', ds_vals.get('id'))
                            insert_link = sa.text(
                                "INSERT INTO project_data_source (id, project_id, data_source_id, data_source_type, is_active, added_at) VALUES (gen_random_uuid(), :project_id, :data_source_id, :data_source_type, :is_active, :added_at)"
                            )
                            conn.execute(insert_link, {"project_id": int(project_id), "data_source_id": ds_vals.get('id'), "data_source_type": data_source_data['type'], "is_active": True, "added_at": datetime.now()})

                        # Update in-memory registry
                        # Persist a copy in the in-memory registry including sample rows if provided
                        self.data_sources[ds_vals.get('id')] = {
                            'id': ds_vals.get('id'),
                            'name': ds_vals.get('name'),
                            'type': ds_vals.get('type'),
                            'format': ds_vals.get('format'),
                            'schema': ds_vals.get('schema'),
                            'data': ds_vals.get('data') or [],
                            'created_at': ds_vals.get('created_at').isoformat() if ds_vals.get('created_at') else None,
                            'updated_at': ds_vals.get('updated_at').isoformat() if ds_vals.get('updated_at') else None,
                            'is_active': True,
                            'organization_id': organization_id,
                            'project_id': project_id
                        }
                        return {
                            'success': True,
                            'data_source': {
                                'id': ds_vals.get('id'),
                                'name': ds_vals.get('name'),
                                'type': ds_vals.get('type'),
                                'organization_id': organization_id,
                                'project_id': project_id
                            },
                            'data_source_id': ds_vals.get('id'),
                            'id': ds_vals.get('id'),
                            'message': 'Data source created (queued sync)'
                        }
                    except Exception as e:
                        logger.exception("Sync queued create failed: %s", e)
                        return {'success': False, 'error': str(e)}

                # Serialize DB writes via the write queue to avoid asyncpg "another operation is in progress" errors
                # Use the sync callable so all DB writes run via the sync engine in a worker thread.
                # Retry enqueue a few times in case of transient connection/loop issues, then fall back to async path.
                import asyncio
                max_enqueue_attempts = 3
                for attempt in range(1, max_enqueue_attempts + 1):
                    try:
                        return await write_queue.enqueue(_do_create_sync)
                    except Exception as enqueue_exc:
                        logger.warning("write_queue.enqueue attempt %d/%d failed: %s", attempt, max_enqueue_attempts, enqueue_exc)
                        if attempt == max_enqueue_attempts:
                            logger.exception("write_queue.enqueue failed after %d attempts, falling back to direct async create", max_enqueue_attempts)
                            return await _do_create()
                        await asyncio.sleep(0.4 * attempt)
            except Exception:
                # If queue not available, fall back to direct async create
                return await _do_create()
        except Exception as e:
            # Log full exception with traceback to help debugging in tests
            logger.exception("❌ Failed to create project data source")
            return {
                'success': False,
                'error': repr(e) or str(e) or 'Unknown error'
            }

    async def get_project_data_source(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_id: str
    ) -> Dict[str, Any]:
        """Get a specific data source for a project"""
        try:
            logger.info(f"🔍 Getting project data source {data_source_id} for project {project_id}")
            
            from app.modules.data.models import DataSource
            from app.modules.projects.models import ProjectDataSource
            from app.db.session import async_session

            async with async_session() as db:
                from sqlalchemy import select, join
                
                query = (
                    select(DataSource)
                    .join(ProjectDataSource, DataSource.id == ProjectDataSource.data_source_id)
                    .where(
                        DataSource.id == data_source_id,
                        ProjectDataSource.project_id == int(project_id),
                        DataSource.is_active == True,
                        ProjectDataSource.is_active == True
                    )
                )
                
                result = await db.execute(query)
                data_source = result.scalar_one_or_none()
                
                if data_source:
                    return {
                        'success': True,
                        'data_source': {
                            'id': data_source.id,
                            'name': data_source.name,
                            'type': data_source.type,
                            'format': data_source.format,
                            'size': data_source.size,
                            'row_count': data_source.row_count,
                            'schema': data_source.schema,
                            'organization_id': organization_id,
                            'project_id': project_id,
                            'created_at': data_source.created_at.isoformat() if data_source.created_at else None,
                            'updated_at': data_source.updated_at.isoformat() if data_source.updated_at else None
                        }
                    }
                else:
                    return {
                        'success': False,
                        'error': 'Data source not found'
                    }
        except Exception as e:
            logger.error(f"❌ Failed to get project data source: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def execute_project_data_source_query(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_id: str, 
        query: str
    ) -> Dict[str, Any]:
        """Execute a query on a project data source"""
        try:
            logger.info(f"🔍 Executing project data source query for {data_source_id}")
            
            # First verify the data source belongs to the project
            data_source_result = await self.get_project_data_source(organization_id, project_id, data_source_id)
            if not data_source_result['success']:
                return data_source_result
            
            # Execute the query using the existing method
            return await self.execute_query_on_source(data_source_id, query)
        except Exception as e:
            logger.error(f"❌ Failed to execute project data source query: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }

    async def get_project_data_source_data(
        self, 
        organization_id: str, 
        project_id: str, 
        data_source_id: str, 
        limit: int = 100
    ) -> Dict[str, Any]:
        """Get data from a project data source"""
        try:
            logger.info(f"📊 Getting project data source data for {data_source_id}")
            
            # First verify the data source belongs to the project
            data_source_result = await self.get_project_data_source(organization_id, project_id, data_source_id)
            if not data_source_result['success']:
                return data_source_result
            
            # Get the data using the existing method
            return await self.get_data_from_source(data_source_id, limit)
        except Exception as e:
            logger.error(f"❌ Failed to get project data source data: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }